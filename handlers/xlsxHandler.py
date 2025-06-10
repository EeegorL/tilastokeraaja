from openpyxl import *;
from utils.utils import sheetColumn;
import os;
from datetime import datetime;

def getFileName():
    date = datetime.today();
    dd = str(date.day) if len(str(date.day)) == 2 else f"0{str(date.day)}";
    mm = str(date.month) if len(str(date.month)) == 2 else f"0{str(date.month)}";
    yyyy = str(date.year);

    date = f"{dd}{mm}{yyyy}";

    return f"/tulokset_{date}.xlsx"

def delXlsx(folder: str) -> None:
    fName = getFileName();
    try:
        os.remove(folder+fName);
    except FileNotFoundError:
        pass;

def doXlsxThings(data, folder: str):
    fName = getFileName();
    
    dummyWorkbook = Workbook();
    if(os.path.isfile(folder+getFileName())):
        dummyWorkbook = load_workbook(folder+getFileName());
    
    sheet = dummyWorkbook["Sheet"];
    column = sheetColumn(data["location"]); # B/C/D/E depending on location
    row = 2;
    dateCounter = 2;

    # header
    sheet["A1"] = data["time"];
    sheet[f"{column}{str(row - 1)}"] = data["location"];

    # data
    for item in data["dates"]:
        date = item;
        dateSplit = date.split("-");
        cell = sheet[f"A{str(dateCounter)}"];
        cell.number_format = "DD.MM.YYYY;@";

        dd = dateSplit[2];
        mm = dateSplit[1];
        yyyy = dateSplit[0];

        sheet[f"A{str(dateCounter)}"] = f"{dd}.{mm}.{yyyy}";

        dateCounter += 1;

    for item in data["data"]:
        sheet[f"{column}{str(row)}"] = item;
        row += 1;
    try:
        dummyWorkbook.save(folder+fName);
        return True;
    except PermissionError:
        return False;

def xlsxIsOpen(folder):
    fName = getFileName();
    if(not os.path.isfile(folder+fName)):
        dummyWorkbook = Workbook();
        try:
            dummyWorkbook.save(folder+fName);
            return False;
        except PermissionError:
            return True;
    else:
        dummyWorkbook = load_workbook(folder+fName);
        try:
            dummyWorkbook.save(folder+fName);
            return False;
        except PermissionError:
            return True;     